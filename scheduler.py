import json
from ortools.sat.python import cp_model
from collections import defaultdict
import os
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import sys

# ------------------------- Налаштування -------------------------
# Ці константи будуть перевизначені на основі вводу користувача в GUI
DEFAULT_SLOTS_PER_DAY = 5 
DAYS = ["Пн", "Вт", "Ср", "Чт", "Пт"]

# ------------------------- Завантаження даних -------------------------
def load_json(path):
    """Завантажує JSON-файл з вказаного шляху."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        # Ця помилка обробляється вище в run_solver_and_generate_reports
        # або якщо якась інша функція намагається завантажити файл напряму
        messagebox.showerror("Помилка завантаження", f"Файл не знайдено: {path}.")
        raise
    except json.JSONDecodeError:
        messagebox.showerror("Помилка JSON", f"Помилка декодування JSON у файлі: {path}. Переконайтеся, що файл має коректний формат JSON.")
        raise

# Клас для представлення однієї лекції (пари) з усіма її атрибутами.
class Lecture:
    """Представляє одну лекцію (пару) з усіма її атрибутами."""
    def __init__(self, group, subject, teacher, count):
        self.group = group
        self.subject = subject
        self.teacher = teacher
        self.count = count # Кількість годин/пар на тиждень для цього предмета
        self.vars = [] # Змінні CP-SAT для слотів і кімнат цієї лекції

def run_solver_and_generate_reports(data_folder, strategy_choice, user_slots_per_day):
    """
    Запускає CP-SAT розв'язувач для генерації розкладу
    та повертає дані розкладу для відображення та збереження.
    """
    # Оновлення глобальних констант на основі вводу користувача
    try:
        SLOTS_PER_DAY = int(user_slots_per_day)
        if SLOTS_PER_DAY <= 0:
            raise ValueError("Кількість пар на день має бути позитивним цілим числом.")
    except ValueError as e:
        return None, None, None, f"Помилка вводу: {e}. Будь ласка, введіть дійсне число для 'Бажана кількість пар на день'."

    TOTAL_SLOTS = len(DAYS) * SLOTS_PER_DAY

    # Перевірка наявності необхідних файлів у папці
    required_files = ["groups.json", "teachers.json", "subjects.json", "rooms.json"]
    missing_files = []
    for filename in required_files:
        if not os.path.exists(os.path.join(data_folder, filename)):
            missing_files.append(filename)
    
    if missing_files:
        error_message = "Відсутні наступні файли у вибраній папці:\n" + "\n".join(missing_files) + "\nБудь ласка, переконайтеся, що всі необхідні JSON файли знаходяться у вказаній папці."
        messagebox.showerror("Помилка вхідних даних", error_message)
        return None, None, None, "Помилка вхідних даних: відсутні файли."


    try:
        # Завантаження даних з файлів
        groups = load_json(os.path.join(data_folder, "groups.json"))
        teachers = load_json(os.path.join(data_folder, "teachers.json"))
        subjects = load_json(os.path.join(data_folder, "subjects.json"))
        rooms = load_json(os.path.join(data_folder, "rooms.json"))
    except Exception as e:
        # Цей блок відловить помилки JSONDecodeError або інші невідомі помилки
        return None, None, None, f"Помилка завантаження даних: {e}"

    # Створення словника для швидкого доступу до типів предметів
    subject_types = {s["name"]: s.get("type", "") for s in subjects}

    # ------------------------- Модель розкладу -------------------------
    model = cp_model.CpModel()

    # Список для зберігання всіх об'єктів Lecture
    lectures = []
    # Діапазони індексів для слотів та кімнат
    slot_indices = list(range(TOTAL_SLOTS))
    room_indices = list(range(len(rooms)))

    # Створення об'єктів Lecture на основі вхідних даних
    for group in groups:
        for subj in group["subjects"]:
            teacher = subj["teacher"]
            name = subj["name"]
            count = subj["hours"]
            # Переконайтеся, що загальна кількість годин для предмета не перевищує TOTAL_SLOTS * кількість груп для цього типу предмета
            if count > TOTAL_SLOTS:
                messagebox.showerror("Помилка вхідних даних", f"Предмет '{name}' для групи '{group['name']}' має {count} годин, що перевищує загальну доступну кількість слотів ({TOTAL_SLOTS}) для однієї групи. Будь ласка, скоригуйте години або кількість пар на день.")
                return None, None, None, "Помилка вхідних даних: години перевищують загальну кількість слотів."
            lectures.append(Lecture(group["name"], name, teacher, count))

    # ------------------------- Змінні -------------------------
    # Список для зберігання всіх лекцій з їхніми змінними
    schedule = []
    # Словники для відстеження слотів за групою та викладачем (для обмежень)
    all_slots_by_group = defaultdict(list)
    all_slots_by_teacher = defaultdict(list)

    # Створення змінних для кожного екземпляра лекції (слот і кімната)
    for lecture in lectures:
        vars_per_lecture = []
        for i in range(lecture.count):
            # Змінна для часового слоту (від 0 до TOTAL_SLOTS - 1)
            slot = model.NewIntVar(0, TOTAL_SLOTS - 1, f"slot_{lecture.group}_{lecture.subject}_{i}")
            # Змінна для кімнати (від 0 до len(rooms) - 1)
            room = model.NewIntVar(0, len(rooms) - 1, f"room_{lecture.group}_{lecture.subject}_{i}")
            vars_per_lecture.append((slot, room))
            all_slots_by_group[lecture.group].append(slot)
            all_slots_by_teacher[lecture.teacher].append(slot)
        lecture.vars = vars_per_lecture
        schedule.append(lecture)

    # ------------------------- Жорсткі обмеження -------------------------
    # Словники для відстеження зайнятих слотів за групою та викладачем
    used_slots = defaultdict(list)
    # Список для кодування пари "слот-кімната" для унікальності
    room_slot_encodings = []
    # Словники для відстеження слотів за групою та викладачем для обмежень AllDifferent
    slot_by_group = defaultdict(list)
    slot_by_teacher = defaultdict(list)
    # Словник для підрахунку пар на день для кожної групи
    day_slot_count_group = defaultdict(lambda: defaultdict(list))

    for lec in schedule:
        for i, (slot, room) in enumerate(lec.vars):
            # Обмеження: одна група не може мати дві пари одночасно
            # Обмеження: один викладач не може вести дві пари одночасно
            # Створення унікального ключа для комбінації (слот, кімната) для групи та викладача
            group_key = model.NewIntVar(0, TOTAL_SLOTS * len(rooms) - 1, f"group_slot_{lec.group}_{i}")
            teacher_key = model.NewIntVar(0, TOTAL_SLOTS * len(rooms) - 1, f"teacher_slot_{lec.teacher}_{i}")
            model.Add(group_key == slot * len(rooms) + room)
            model.Add(teacher_key == slot * len(rooms) + room)

            used_slots[f"group:{lec.group}"].append(group_key)
            used_slots[f"teacher:{lec.teacher}"].append(teacher_key)

            # Обмеження: тип аудиторії повинен відповідати типу предмета
            subject_type = subject_types.get(lec.subject, "")
            for room_index, r in enumerate(rooms):
                room_type = r.get("type", "")
                if room_type != subject_type and subject_type != "": # Якщо тип предмета вказано
                    # Якщо тип кімнати не відповідає типу предмета, ця кімната не може бути призначена
                    model.Add(room != room_index).OnlyEnforceIf(
                        model.NewBoolVar(f"room_type_mismatch_{lec.group}_{lec.subject}_{i}_{room_index}")
                    )
            
            # Обмеження: одна кімната може бути зайнята лише однією парою в один слот
            room_slot_key = model.NewIntVar(0, TOTAL_SLOTS * len(rooms) - 1, f"room_slot_{lec.group}_{i}")
            model.Add(room_slot_key == slot * len(rooms) + room)
            room_slot_encodings.append(room_slot_key)

            # Зберігаємо змінні слотів окремо для обмежень AllDifferent
            slot_by_group[lec.group].append(slot)
            slot_by_teacher[lec.teacher].append(slot)

            # Обмеження: не більше SLOTS_PER_DAY пар на день для кожної групи
            # Створення булевих змінних, які вказують, чи лекція припадає на певний день
            for d in range(len(DAYS)):
                in_day = model.NewBoolVar(f"is_{lec.group}_{lec.subject}_{i}_day{d}")
                # Створення булевих змінних для кожного з лінійних виразів
                is_slot_ge_lower = model.NewBoolVar(f"slot_ge_lower_{lec.group}_{lec.subject}_{i}_day{d}")
                is_slot_lt_upper = model.NewBoolVar(f"slot_lt_upper_{lec.group}_{lec.subject}_{i}_day{d}")

                # Прив'язка булевих змінних до лінійних виразів
                model.Add(slot >= d * SLOTS_PER_DAY).OnlyEnforceIf(is_slot_ge_lower)
                model.Add(slot < (d + 1) * SLOTS_PER_DAY).OnlyEnforceIf(is_slot_lt_upper)

                # Тепер використовуємо булеві змінні в AddBoolAnd та OnlyEnforceIf
                model.AddBoolAnd([is_slot_ge_lower, is_slot_lt_upper]).OnlyEnforceIf(in_day)
                model.AddBoolOr([is_slot_ge_lower.Not(), is_slot_lt_upper.Not()]).OnlyEnforceIf(in_day.Not())
                
                day_slot_count_group[lec.group][d].append(in_day)

    # Застосування обмежень AllDifferent:
    # Кожна комбінація (група/викладач, слот, кімната) повинна бути унікальною
    for key, keys in used_slots.items():
        if keys: # Тільки якщо є змінні для застосування AllDifferent
            model.AddAllDifferent(keys)

    # Кожна комбінація (слот, кімната) повинна бути унікальною (одна кімната - одна пара)
    if room_slot_encodings:
        model.AddAllDifferent(room_slot_encodings)

    # Кожна група не може мати дві пари в один і той же слот
    for group, slots in slot_by_group.items():
        if slots: # Тільки якщо є змінні для застосування AllDifferent
            model.AddAllDifferent(slots)

    # Кожен викладач не може вести дві пари в один і той же слот
    for teacher, slots in slot_by_teacher.items():
        if slots: # Тільки якщо є змінні для застосування AllDifferent
            model.AddAllDifferent(slots)

    # Обмеження на максимальну кількість пар на день для групи
    for group, days in day_slot_count_group.items():
        for d, bool_vars in days.items():
            model.Add(sum(bool_vars) <= SLOTS_PER_DAY)

    # --- М'яке обмеження: мінімізація вікон у розкладі ---

    # Створення булевих змінних, що вказують, чи зайнятий певний слот для групи/викладача в конкретний день
    # group_day_slot_occupied[group_name][day_index][slot_in_day_index]
    group_names = [group["name"] for group in groups]
    teacher_names = [teacher["name"] for teacher in teachers]

    group_day_slot_occupied = defaultdict(lambda: defaultdict(lambda: [
        model.NewBoolVar(f'group_occupied_{g}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)
    ]))

    # teacher_day_slot_occupied[teacher_name][day_index][slot_in_day_index]
    teacher_day_slot_occupied = defaultdict(lambda: defaultdict(lambda: [
        model.NewBoolVar(f'teacher_occupied_{t}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)
    ]))


    # Зв'язування булевих змінних зайнятості з фактичними призначеннями лекцій
    for g in group_names:
        for d_idx in range(len(DAYS)):
            for s_idx in range(SLOTS_PER_DAY):
                global_slot_idx = d_idx * SLOTS_PER_DAY + s_idx
                
                # Для груп: збираємо всі літерали, які вказують на те, що лекція цієї групи знаходиться в даному глобальному слоті
                literals_for_group_slot = []
                for lec in schedule:
                    if lec.group == g:
                        for slot_var, _ in lec.vars:
                            # Створення булевої змінної, яка є істиною, якщо ця конкретна лекція знаходиться в global_slot_idx
                            is_this_lec_at_this_global_slot = model.NewBoolVar(f"is_lec_group_{g}_{lec.subject}_instance_at_slot{global_slot_idx}")
                            model.Add(slot_var == global_slot_idx).OnlyEnforceIf(is_this_lec_at_this_global_slot)
                            model.Add(slot_var != global_slot_idx).OnlyEnforceIf(is_this_lec_at_this_global_slot.Not())
                            literals_for_group_slot.append(is_this_lec_at_this_global_slot)
                
                # Якщо є лекції для цієї групи, то group_day_slot_occupied[g][d_idx][s_idx] є істиною, якщо хоча б одна з них у цьому слоті
                if literals_for_group_slot:
                    model.AddBoolOr(literals_for_group_slot).OnlyEnforceIf(group_day_slot_occupied[g][d_idx][s_idx])
                    model.AddBoolAnd([lit.Not() for lit in literals_for_group_slot]).OnlyEnforceIf(group_day_slot_occupied[g][d_idx][s_idx].Not())
                else: # Якщо для цієї групи немає лекцій, які б потрапляли в цей слот, то він точно не зайнятий
                    model.Add(group_day_slot_occupied[g][d_idx][s_idx] == False)


    for t in teacher_names:
        for d_idx in range(len(DAYS)):
            for s_idx in range(SLOTS_PER_DAY):
                global_slot_idx = d_idx * SLOTS_PER_DAY + s_idx
                
                # Для викладачів: збираємо всі літерали, які вказують на те, що лекція цього викладача знаходиться в даному глобальному слоті
                literals_for_teacher_slot = []
                for lec in schedule:
                    if lec.teacher == t:
                        for slot_var, _ in lec.vars:
                            is_this_lec_at_this_global_slot = model.NewBoolVar(f"is_lec_teacher_{t}_{lec.subject}_instance_at_slot{global_slot_idx}")
                            model.Add(slot_var == global_slot_idx).OnlyEnforceIf(is_this_lec_at_this_global_slot)
                            model.Add(slot_var != global_slot_idx).OnlyEnforceIf(is_this_lec_at_this_global_slot.Not())
                            literals_for_teacher_slot.append(is_this_lec_at_this_global_slot)

                if literals_for_teacher_slot:
                    model.AddBoolOr(literals_for_teacher_slot).OnlyEnforceIf(teacher_day_slot_occupied[t][d_idx][s_idx])
                    model.AddBoolAnd([lit.Not() for lit in literals_for_teacher_slot]).OnlyEnforceIf(teacher_day_slot_occupied[t][d_idx][s_idx].Not())
                else: # Якщо для цього викладача немає лекцій, які б потрапляли в цей слот, то він точно не зайнятий
                    model.Add(teacher_day_slot_occupied[t][d_idx][s_idx] == False)

    # Змінна для підрахунку загальної кількості вікон
    total_windows_count = model.NewIntVar(0, TOTAL_SLOTS * (len(groups) + len(teachers)), 'total_windows_count')
    all_window_literals = [] # Список для збору всіх булевих змінних "вікон"

    # Розрахунок вікон для груп (оновлена логіка)
    for g in group_names:
        for d_idx in range(len(DAYS)):
            # Створюємо булеві змінні для перевірки наявності зайнятих слотів до/після поточного
            has_prev_occupied_slots = [model.NewBoolVar(f'group_prev_occ_{g}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)]
            has_next_occupied_slots = [model.NewBoolVar(f'group_next_occ_{g}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)]

            for s_idx in range(SLOTS_PER_DAY):
                # Чи є хоча б один зайнятий слот до поточного (s_idx)?
                if s_idx > 0:
                    model.AddBoolOr([group_day_slot_occupied[g][d_idx][i] for i in range(s_idx)]).OnlyEnforceIf(has_prev_occupied_slots[s_idx])
                    model.AddBoolAnd([group_day_slot_occupied[g][d_idx][i].Not() for i in range(s_idx)]).OnlyEnforceIf(has_prev_occupied_slots[s_idx].Not())
                else:
                    model.Add(has_prev_occupied_slots[s_idx] == False) # Для першого слота немає попередніх

                # Чи є хоча б один зайнятий слот після поточного (s_idx)?
                if s_idx < SLOTS_PER_DAY - 1:
                    model.AddBoolOr([group_day_slot_occupied[g][d_idx][i] for i in range(s_idx + 1, SLOTS_PER_DAY)]).OnlyEnforceIf(has_next_occupied_slots[s_idx])
                    model.AddBoolAnd([group_day_slot_occupied[g][d_idx][i].Not() for i in range(s_idx + 1, SLOTS_PER_DAY)]).OnlyEnforceIf(has_next_occupied_slots[s_idx].Not())
                else:
                    model.Add(has_next_occupied_slots[s_idx] == False) # Для останнього слота немає наступних
                
                # Перевіряємо, чи поточний слот є "вікном"
                current_slot_occupied_literal = group_day_slot_occupied[g][d_idx][s_idx]
                is_group_window_slot = model.NewBoolVar(f"is_group_window_slot_{g}_day{d_idx}_slot{s_idx}")
                
                # Слот є вікном, якщо він вільний І є заняття до нього І є заняття після нього
                model.AddBoolAnd([current_slot_occupied_literal.Not(),
                                  has_prev_occupied_slots[s_idx],
                                  has_next_occupied_slots[s_idx]]).OnlyEnforceIf(is_group_window_slot)
                all_window_literals.append(is_group_window_slot)

    # Розрахунок вікон для викладачів (аналогічна логіка)
    for t in teacher_names:
        for d_idx in range(len(DAYS)):
            # Створюємо булеві змінні для перевірки наявності зайнятих слотів до/після поточного
            has_prev_occupied_slots = [model.NewBoolVar(f'teacher_prev_occ_{t}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)]
            has_next_occupied_slots = [model.NewBoolVar(f'teacher_next_occ_{t}_{d_idx}_{s_idx}') for s_idx in range(SLOTS_PER_DAY)]

            for s_idx in range(SLOTS_PER_DAY):
                # Чи є хоча б один зайнятий слот до поточного (s_idx)?
                if s_idx > 0:
                    model.AddBoolOr([teacher_day_slot_occupied[t][d_idx][i] for i in range(s_idx)]).OnlyEnforceIf(has_prev_occupied_slots[s_idx])
                    model.AddBoolAnd([teacher_day_slot_occupied[t][d_idx][i].Not() for i in range(s_idx)]).OnlyEnforceIf(has_prev_occupied_slots[s_idx].Not())
                else:
                    model.Add(has_prev_occupied_slots[s_idx] == False)

                # Чи є хоча б один зайнятий слот після поточного (s_idx)?
                if s_idx < SLOTS_PER_DAY - 1:
                    model.AddBoolOr([teacher_day_slot_occupied[t][d_idx][i] for i in range(s_idx + 1, SLOTS_PER_DAY)]).OnlyEnforceIf(has_next_occupied_slots[s_idx])
                    model.AddBoolAnd([teacher_day_slot_occupied[t][d_idx][i].Not() for i in range(s_idx + 1, SLOTS_PER_DAY)]).OnlyEnforceIf(has_next_occupied_slots[s_idx].Not())
                else:
                    model.Add(has_next_occupied_slots[s_idx] == False)

                # Перевіряємо, чи поточний слот є "вікном"
                current_slot_occupied_literal = teacher_day_slot_occupied[t][d_idx][s_idx]
                is_teacher_window_slot = model.NewBoolVar(f"is_teacher_window_slot_{t}_day{d_idx}_slot{s_idx}")

                # Слот є вікном, якщо він вільний І є заняття до нього І є заняття після нього
                model.AddBoolAnd([current_slot_occupied_literal.Not(),
                                  has_prev_occupied_slots[s_idx],
                                  has_next_occupied_slots[s_idx]]).OnlyEnforceIf(is_teacher_window_slot)
                all_window_literals.append(is_teacher_window_slot)

    # Додаємо суму всіх булевих змінних "вікон" до моделі
    model.Add(total_windows_count == sum(all_window_literals))

    # ------------------------- Розв’язання -------------------------
    solver = cp_model.CpSolver()
    # Налаштування стратегії пошуку за аргументом командного рядка
    if strategy_choice == "random":
        solver.parameters.random_seed = 42
        solver.parameters.search_branching = cp_model.PORTFOLIO_SEARCH
    elif strategy_choice == "default":
        # Явна установка default, хоча це і так поведінка за замовчуванням
        solver.parameters.search_branching = cp_model.FIXED_SEARCH

    # Встановлення функції цілі: мінімізувати загальну кількість вікон
    model.Minimize(total_windows_count)

    # Запуск розв'язувача
    status = solver.Solve(model)

    # ------------------------- Обробка результатів -------------------------
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # Словники для зберігання розкладу для груп та викладачів
        timetable = defaultdict(lambda: defaultdict(list))
        timetable_teachers = defaultdict(lambda: defaultdict(list))

        for lec in schedule:
            for i, (slot, room) in enumerate(lec.vars):
                time_slot = solver.Value(slot)
                day_index = time_slot // SLOTS_PER_DAY
                pair = time_slot % SLOTS_PER_DAY + 1
                day = DAYS[day_index]
                room_name = rooms[solver.Value(room)]["name"]

                # Зберігаємо окремі компоненти даних
                timetable[lec.group][day].append((pair, lec.subject, lec.teacher, room_name))
                timetable_teachers[lec.teacher][day].append((pair, lec.subject, lec.group, room_name))

        # Створення директорії для експорту, якщо вона не існує
        export_folder = os.path.join(os.getcwd(), "export")
        os.makedirs(export_folder, exist_ok=True)

        # Excel для груп
        wb = Workbook()
        # Видалення стандартного аркуша 'Sheet', якщо він був створений
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        for group, days in timetable.items():
            # Створення аркуша для кожної групи, обмежуючи назву до 31 символу
            ws = wb.create_sheet(title=group[:31])
            # Заголовок стовпців
            ws.append(["День", "Пара", "Предмет", "Викладач", "Аудиторія"])
            for day in DAYS:
                entries = days.get(day, [])
                # Сортування записів за номером пари
                for entry in sorted(entries):
                    # Розпаковуємо дані безпосередньо з кортежу
                    pair, subject_name, teacher_name, auditorium_name = entry
                    ws.append([day, pair, subject_name, teacher_name, auditorium_name])
        # Збереження файлу розкладу для груп
        schedule_filepath = os.path.join(export_folder, "schedule.xlsx")
        wb.save(schedule_filepath)

        # Excel для викладачів
        wb_t = Workbook()
        # Видалення стандартного аркуша 'Sheet', якщо він був створений
        if 'Sheet' in wb_t.sheetnames:
            del wb_t['Sheet']
        for teacher_name, days in timetable_teachers.items():
            # Створення аркуша для кожного викладача, обмежуючи назву до 31 символу
            ws = wb_t.create_sheet(title=teacher_name[:31])
            # Заголовок стовпців
            ws.append(["День", "Пара", "Предмет", "Група", "Аудиторія"])
            for day in DAYS:
                entries = days.get(day, [])
                # Сортування записів за номером пари
                for entry in sorted(entries):
                    # Розпаковуємо дані безпосередньо з кортежу
                    pair, subject_name, group_name, auditorium_name = entry
                    ws.append([day, pair, subject_name, group_name, auditorium_name])
        # Збереження файлу розкладу для викладачів
        teachers_schedule_filepath = os.path.join(export_folder, "teachers_schedule.xlsx")
        wb_t.save(teachers_schedule_filepath)

        # Формування детального звіту про вікна
        report_text = ["\n--- Детальний звіт про вікна ---"]
        calculated_windows_count_debugger = 0

        # Звіт для груп
        for g in group_names:
            for d_idx, day in enumerate(DAYS):
                occupied_slots_representation = []
                for s_idx in range(SLOTS_PER_DAY):
                    if solver.Value(group_day_slot_occupied[g][d_idx][s_idx]):
                        occupied_slots_representation.append("X")
                    else:
                        occupied_slots_representation.append("O")
                
                first_occupied = -1
                last_occupied = -1
                for i, status in enumerate(occupied_slots_representation):
                    if status == "X":
                        if first_occupied == -1:
                            first_occupied = i
                        last_occupied = i
                
                windows_for_this_day = 0
                if first_occupied != -1:
                    for s_idx in range(first_occupied + 1, last_occupied):
                        if occupied_slots_representation[s_idx] == "O":
                            windows_for_this_day += 1
                            calculated_windows_count_debugger += 1
                    report_text.append(f"Група {g}, {day}: {windows_for_this_day} вікон. Розклад: {''.join(occupied_slots_representation)}")
                else:
                    report_text.append(f"Група {g}, {day}: Немає занять.")

        # Звіт для викладачів
        for t in teacher_names:
            for d_idx, day in enumerate(DAYS):
                occupied_slots_representation = []
                for s_idx in range(SLOTS_PER_DAY):
                    if solver.Value(teacher_day_slot_occupied[t][d_idx][s_idx]):
                        occupied_slots_representation.append("X")
                    else:
                        occupied_slots_representation.append("O")
                
                first_occupied = -1
                last_occupied = -1
                for i, status in enumerate(occupied_slots_representation):
                    if status == "X":
                        if first_occupied == -1:
                            first_occupied = i
                        last_occupied = i
                
                windows_for_this_day = 0
                if first_occupied != -1:
                    for s_idx in range(first_occupied + 1, last_occupied):
                        if occupied_slots_representation[s_idx] == "O":
                            windows_for_this_day += 1
                            calculated_windows_count_debugger += 1
                    report_text.append(f"Викладач {t}, {day}: {windows_for_this_day} вікон. Розклад: {''.join(occupied_slots_representation)}")
                else:
                    report_text.append(f"Викладач {t}, {day}: Немає занять.")
        
        report_text.append(f"\n📊 Загальна кількість вікон у розкладі (за цільовою функцією): {int(solver.ObjectiveValue())}")
        report_text.append(f"Підраховано вікон (для перевірки у звіті): {calculated_windows_count_debugger}")
        
        if solver.ObjectiveValue() == 0:
            report_text.append("\n🎉 Оптимальне рішення знайдено: розклад не містить вікон між заняттями.")
        else:
            report_text.append(f"\n💡 Оптимальне рішення знайдено. Залишилося {int(solver.ObjectiveValue())} вікон, яких неможливо уникнути через жорсткі обмеження.")
        
        return timetable, timetable_teachers, "\n".join(report_text), "Розклад успішно згенеровано!"

    else:
        # Випадок, коли рішення не знайдено
        conflict_report_text = (
            "❌ Не вдалося знайти допустиме рішення. Перевірте конфлікти у вхідних даних.\n\n"
            "📌 Можливі причини:\n"
            "- Група перевантажена (занадто багато пар на тиждень)\n"
            "- Аудиторій недостатньо або неправильного типу\n"
            "- Один викладач закріплений за занадто багатьма групами\n"
            "- Всі групи мають пари одночасно, а кімнат не вистачає\n"
            "\n🔎 Перевірте файли у папці data/: groups.json, teachers.json, rooms.json, subjects.json\n"
        )
        export_folder = os.path.join(os.getcwd(), "export")
        os.makedirs(export_folder, exist_ok=True)
        with open(os.path.join(export_folder, "conflict_report.txt"), "w", encoding="utf-8") as f:
            f.write(conflict_report_text)
        return None, None, conflict_report_text, "❌ Не вдалося знайти допустиме рішення. Дивіться 'export/conflict_report.txt'"


class ScheduleApp:
    def __init__(self, master):
        self.master = master
        master.title("Автоматичне складання розкладу")
        master.geometry("1000x700") # Початковий розмір вікна

        self.data_folder = tk.StringVar(value="")
        self.strategy_choice = tk.StringVar(value="default")
        self.user_slots_per_day = tk.StringVar(value=str(DEFAULT_SLOTS_PER_DAY)) # Нова змінна для вводу користувача

        self.create_widgets()

    def create_widgets(self):
        # Фрейм для вибору шляху до даних та стратегії
        control_frame = ttk.Frame(self.master, padding="10")
        control_frame.pack(fill=tk.X)

        ttk.Label(control_frame, text="Папка з даними:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.path_entry = ttk.Entry(control_frame, textvariable=self.data_folder, width=50)
        self.path_entry.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        self.path_entry.bind("<Button-1>", lambda e: self.browse_folder()) # Прив'язка кліку для вибору папки
        
        browse_button = ttk.Button(control_frame, text="Обрати папку", command=self.browse_folder)
        browse_button.grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(control_frame, text="Стратегія пошуку:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(control_frame, text="За замовчуванням", variable=self.strategy_choice, value="default").grid(row=1, column=1, sticky=tk.W)
        ttk.Radiobutton(control_frame, text="Випадкова (Portfolio Search)", variable=self.strategy_choice, value="random").grid(row=1, column=2, sticky=tk.W)

        # Новий ввід для бажаної кількості пар на день
        ttk.Label(control_frame, text="Бажана кількість пар на день:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.slots_per_day_entry = ttk.Entry(control_frame, textvariable=self.user_slots_per_day, width=10)
        self.slots_per_day_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)

        generate_button = ttk.Button(control_frame, text="Згенерувати розклад", command=self.generate_schedule)
        generate_button.grid(row=3, column=0, columnspan=3, pady=10)

        self.status_label = ttk.Label(self.master, text="Очікування...", foreground="blue")
        self.status_label.pack(pady=5)

        # Ноутбук для відображення розкладів
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Вкладка для детального звіту (вікна)
        self.report_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.report_tab, text="Звіт про вікна")
        self.report_text = tk.Text(self.report_tab, wrap=tk.WORD, state=tk.DISABLED, width=80, height=20)
        self.report_text.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        # Додавання смуги прокрутки до вкладки звіту
        report_scrollbar = ttk.Scrollbar(self.report_text, orient=tk.VERTICAL, command=self.report_text.yview)
        report_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.report_text.config(yscrollcommand=report_scrollbar.set)

        # Кнопки завантаження
        download_frame = ttk.Frame(self.master, padding="10")
        download_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.download_group_btn = ttk.Button(download_frame, text="Завантажити розклад груп (.xlsx)", command=lambda: self.download_file("schedule.xlsx"))
        self.download_group_btn.pack(side=tk.LEFT, padx=10, pady=5)
        self.download_group_btn["state"] = tk.DISABLED

        self.download_teacher_btn = ttk.Button(download_frame, text="Завантажити розклад викладачів (.xlsx)", command=lambda: self.download_file("teachers_schedule.xlsx"))
        self.download_teacher_btn.pack(side=tk.LEFT, padx=10, pady=5)
        self.download_teacher_btn["state"] = tk.DISABLED

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.data_folder.set(folder_selected)
            self.status_label.config(text=f"Вибрано папку: {folder_selected}", foreground="black")

    def generate_schedule(self):
        data_path = self.data_folder.get()
        if not data_path:
            messagebox.showwarning("Помилка", "Будь ласка, оберіть папку з даними.")
            return

        self.status_label.config(text="Генеруємо розклад...", foreground="orange")
        self.master.update_idletasks() # Оновити GUI негайно

        # Очистити попередні вкладки
        for tab in self.notebook.tabs():
            if tab != self.report_tab.winfo_id(): # Не видаляти вкладку звіту
                self.notebook.forget(tab)
        
        # Вимкнути кнопки під час генерації
        self.download_group_btn["state"] = tk.DISABLED
        self.download_teacher_btn["state"] = tk.DISABLED
        
        # Отримати бажану кількість слотів на день від користувача
        user_slots_per_day_value = self.user_slots_per_day.get()

        # Запустити логіку розв'язувача
        timetable, timetable_teachers, report_text, status_message = run_solver_and_generate_reports(
            data_path, self.strategy_choice.get(), user_slots_per_day_value
        )
        
        self.status_label.config(text=status_message, 
                                 foreground="green" if timetable else "red")
        
        # Оновити вкладку звіту
        self.report_text.config(state=tk.NORMAL)
        self.report_text.delete(1.0, tk.END)
        self.report_text.insert(tk.END, report_text)
        self.report_text.config(state=tk.DISABLED)

        if timetable:
            # Відобразити розклади груп
            for group_name in sorted(timetable.keys()):
                group_tab = ttk.Frame(self.notebook)
                self.notebook.add(group_tab, text=group_name[:31]) # Обрізати назву вкладки
                
                group_schedule_text = tk.Text(group_tab, wrap=tk.WORD, state=tk.DISABLED)
                group_schedule_text.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

                # Додати смугу прокрутки
                group_scrollbar = ttk.Scrollbar(group_schedule_text, orient=tk.VERTICAL, command=group_schedule_text.yview)
                group_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                group_schedule_text.config(yscrollcommand=group_scrollbar.set)

                self.display_schedule_in_text(group_schedule_text, timetable[group_name], "Група", group_name)

            # Відобразити розклади викладачів
            for teacher_name in sorted(timetable_teachers.keys()):
                teacher_tab = ttk.Frame(self.notebook)
                self.notebook.add(teacher_tab, text=teacher_name[:31]) # Обрізати назву вкладки

                teacher_schedule_text = tk.Text(teacher_tab, wrap=tk.WORD, state=tk.DISABLED)
                teacher_schedule_text.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

                # Додати смугу прокрутки
                teacher_scrollbar = ttk.Scrollbar(teacher_schedule_text, orient=tk.VERTICAL, command=teacher_schedule_text.yview)
                teacher_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                teacher_schedule_text.config(yscrollcommand=teacher_scrollbar.set)

                self.display_schedule_in_text(teacher_schedule_text, timetable_teachers[teacher_name], "Викладач", teacher_name)
            
            self.download_group_btn["state"] = tk.NORMAL
            self.download_teacher_btn["state"] = tk.NORMAL
            
            # Вибрати першу вкладку розкладу
            if self.notebook.tabs():
                self.notebook.select(self.notebook.tabs()[1]) # Вибрати першу фактичну вкладку розкладу (після звіту)

    def display_schedule_in_text(self, text_widget, schedule_data, entity_type, entity_name):
        text_widget.config(state=tk.NORMAL)
        text_widget.delete(1.0, tk.END)

        text_widget.insert(tk.END, f"Розклад для {entity_type} {entity_name}:\n\n")
        
        # Налаштувати тег для жирного тексту
        text_widget.tag_configure("bold", font=("TkDefaultFont", 10, "bold"))

        col_widths = {"Пара": 6, "Предмет": 30, "Викладач/Група": 20, "Аудиторія": 15}
        
        header_line = f"{'Пара':<{col_widths['Пара']}} {'Предмет':<{col_widths['Предмет']}} {'Викладач/Група':<{col_widths['Викладач/Група']}} {'Аудиторія':<{col_widths['Аудиторія']}}\n"
        header_separator = "-" * (sum(col_widths.values()) + len(col_widths)*2) + "\n"

        for day in DAYS:
            entries = schedule_data.get(day, [])
            
            # Додати помітний роздільник та заголовок дня для кожного блоку
            text_widget.insert(tk.END, "=" * 70 + "\n")
            text_widget.insert(tk.END, f"  {day}  \n", "bold")
            text_widget.insert(tk.END, "=" * 70 + "\n")

            if entries:
                # Додати підзаголовок для записів дня
                text_widget.insert(tk.END, header_line)
                text_widget.insert(tk.END, header_separator) 

                for entry in sorted(entries):
                    pair, subject, third_col_value, room = entry # third_col_value is teacher for groups, group for teachers
                    line = f"{pair:<{col_widths['Пара']}} {subject:<{col_widths['Предмет']}} {third_col_value:<{col_widths['Викладач/Група']}} {room:<{col_widths['Аудиторія']}}\n"
                    text_widget.insert(tk.END, line)
            else:
                text_widget.insert(tk.END, f"  Немає занять на {day}.\n")
            text_widget.insert(tk.END, "\n") # Додати новий рядок після блоку кожного дня

        text_widget.config(state=tk.DISABLED)

    def download_file(self, filename):
        export_folder = os.path.join(os.getcwd(), "export")
        filepath = os.path.join(export_folder, filename)
        
        if not os.path.exists(filepath):
            messagebox.showerror("Помилка", f"Файл {filename} не знайдено. Можливо, розклад не був згенерований або стався збій.")
            return

        try:
            # Використання filedialog.asksaveasfilename, щоб користувач міг вибрати місце збереження
            save_path = filedialog.asksaveasfilename(
                initialdir=os.path.expanduser("~"), # Початок у домашній директорії користувача
                initialfile=filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if save_path:
                import shutil
                shutil.copy(filepath, save_path)
                messagebox.showinfo("Успіх", f"Файл '{filename}' успішно збережено за шляхом:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Помилка завантаження", f"Не вдалося зберегти файл '{filename}':\n{e}")

if __name__ == "__main__":
    # Створення папки 'data' та порожніх JSON-файлів, якщо вони не існують
    # Це потрібно, щоб програма могла запуститись з початковими даними
    # і користувач міг бачити, як мають виглядати файли.
    current_dir = os.getcwd()
    data_dir = os.path.join(current_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    # Функція для створення порожнього JSON-файлу
    def create_empty_json(filepath, default_content):
        if not os.path.exists(filepath):
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(default_content, f, ensure_ascii=False, indent=4)
            print(f"Створено порожній файл: {filepath}")

    # Створення файлів з мінімальним прикладом
    create_empty_json(os.path.join(data_dir, "groups.json"), [
        {"name": "Група_1", "subjects": [{"name": "Математика", "teacher": "Петров", "hours": 2}]}
    ])
    create_empty_json(os.path.join(data_dir, "teachers.json"), [
        {"name": "Петров"}
    ])
    create_empty_json(os.path.join(data_dir, "subjects.json"), [
        {"name": "Математика", "type": "лекція"},
        {"name": "Фізика", "type": "практика"}
    ])
    create_empty_json(os.path.join(data_dir, "rooms.json"), [
        {"name": "Ауд_101", "type": "лекція"},
        {"name": "Лаб_203", "type": "практика"}
    ])
    print(f"Переконайтеся, що папка 'data' у вашій поточній директорії містить файли 'groups.json', 'teachers.json', 'subjects.json', 'rooms.json'.\n"
          f"Приклади файлів були створені автоматично за шляхом: {data_dir}")

    root = tk.Tk()
    app = ScheduleApp(root)
    root.mainloop()
